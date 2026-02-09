#!/usr/bin/env python3
"""
Convert IEA Critical Minerals Data Explorer Excel to JSON.

Reads CM_Data_Explorer.xlsx and produces data/iea.json for the dashboard.
Run: python3 scripts/convert-iea-data.py [path-to-xlsx]
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip3 install openpyxl")
    sys.exit(1)

XLSX_DEFAULT = Path.home() / "Downloads" / "CM_Data_Explorer.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "data"

# Column indices (1-based) for the common scenario layout (Sheets 1, 3.1, 3.2, 4.x)
# Col B=2024, D-H=STEPS(2030-2050), J-N=APS(2030-2050), P-T=NZE(2030-2050)
SCENARIO_COLS = {
    "2024": 2,
    "STEPS": {2030: 4, 2035: 5, 2040: 6, 2045: 7, 2050: 8},
    "APS":  {2030: 10, 2035: 11, 2040: 12, 2045: 13, 2050: 14},
    "NZE":  {2030: 16, 2035: 17, 2040: 18, 2045: 19, 2050: 20},
}

# Sheet 2 supply columns (mining: cols 1-5, refining: cols 7-11)
SUPPLY_MINING_COLS = {2024: 2, 2030: 3, 2035: 4, 2040: 5}
SUPPLY_REFINING_COLS = {2024: 8, 2030: 9, 2035: 10, 2040: 11}

# Minerals in Sheet 1 (total demand) - order as they appear
SHEET1_MINERALS = [
    "Copper", "Cobalt", "Lithium", "Nickel",
    "Magnet rare earth elements",
    "Graphite (all grades: natural and synthetic)",
]

# Technology sectors in Sheet 1 per mineral
SHEET1_SECTORS = {
    "Copper": ["Solar PV", "Wind", "Other low emissions power generation",
               "Electric vehicles", "Grid battery storage", "Electricity networks",
               "Hydrogen technologies"],
    "Cobalt": ["Low emissions power generation", "Electric vehicles",
               "Grid battery storage", "Hydrogen technologies"],
    "Lithium": ["Electric vehicles", "Grid battery storage"],
    "Nickel": ["Solar PV", "Wind", "Other low emissions power generation",
               "Electric vehicles", "Grid battery storage", "Hydrogen technologies"],
    "Magnet rare earth elements": ["Wind", "Electric vehicles"],
    "Graphite (all grades: natural and synthetic)": ["Electric vehicles", "Grid battery storage"],
}

# Map IEA mineral names to dashboard-friendly keys
IEA_NAME_MAP = {
    "Magnet rare earth elements": "Rare Earth Elements",
    "Graphite (all grades: natural and synthetic)": "Graphite",
    "Battery-grade graphite": "Battery-grade graphite",
    "PGMs (other than iridum)": "PGMs",
    "Total rare earth elements": "Total REE",
}

# Supply mineral header patterns -> clean name
SUPPLY_MINERALS = {
    "Copper": "Copper",
    "Cobalt": "Cobalt",
    "Lithium": "Lithium",
    "Nickel": "Nickel",
    "Graphite": "Graphite",
    "Magnet rare earth elements": "Rare Earth Elements",
}

# Technology sheets: sheet name -> tech display name
TECH_SHEETS = {
    "4.1 Solar PV": "Solar PV",
    "4.2 Wind": "Wind",
    "4.3 EV": "Electric vehicles",
    "4.4 Battery storage": "Grid battery storage",
    "4.5 Electricity networks": "Electricity networks",
    "4.6 Hydrogen": "Hydrogen technologies",
}


def clean_name(name):
    """Map IEA name to dashboard name."""
    return IEA_NAME_MAP.get(name, name)


def read_scenario_row(ws, row):
    """Read a data row with the standard scenario layout. Returns dict."""
    val_2024 = ws.cell(row, SCENARIO_COLS["2024"]).value
    result = {"2024": safe_num(val_2024)}
    for scenario in ("STEPS", "APS", "NZE"):
        sc = {}
        for year, col in SCENARIO_COLS[scenario].items():
            v = ws.cell(row, col).value
            sc[str(year)] = safe_num(v)
        result[scenario] = sc
    return result


def safe_num(v):
    """Convert to float or None."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_sheet1_total_demand(wb):
    """Parse Sheet 1: Total demand for key minerals."""
    ws = wb["1 Total demand for key minerals"]
    result = {}

    current_mineral = None
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()

        # Check if this is a mineral header
        if label in SHEET1_MINERALS:
            current_mineral = clean_name(label)
            result[current_mineral] = {"sectors": {}}
            continue

        if current_mineral is None:
            continue

        # Known aggregate rows
        if label == "Total clean technologies":
            result[current_mineral]["totalClean"] = read_scenario_row(ws, r)
        elif label == "Other uses":
            result[current_mineral]["otherUses"] = read_scenario_row(ws, r)
        elif label == "Total demand":
            result[current_mineral]["totalDemand"] = read_scenario_row(ws, r)
        elif label.startswith("Share of clean technologies"):
            result[current_mineral]["cleanShare"] = read_scenario_row(ws, r)
        elif label in SHEET1_SECTORS.get(
            [k for k in SHEET1_MINERALS if clean_name(k) == current_mineral][0], []
        ):
            result[current_mineral]["sectors"][label] = read_scenario_row(ws, r)

    return result


def parse_sheet2_supply(wb):
    """Parse Sheet 2: Total supply for key minerals."""
    ws = wb["2 Total supply for key minerals"]
    result = {}

    current_mineral = None
    current_side = None  # 'mining' or 'refining'

    for r in range(1, ws.max_row + 1):
        label_col1 = ws.cell(r, 1).value
        label_col7 = ws.cell(r, 7).value

        # Check for mineral section headers
        if label_col1 and " - " in str(label_col1):
            parts = str(label_col1).split(" - ", 1)
            mineral_key = parts[0].strip()
            for k, v in SUPPLY_MINERALS.items():
                if mineral_key.startswith(k):
                    current_mineral = v
                    break
            if current_mineral not in result:
                result[current_mineral] = {"mining": {"countries": {}}, "refining": {"countries": {}}}
            current_side = "mining"

            # Also check refining header on same row
            if label_col7 and " - " in str(label_col7):
                pass  # refining data in cols 7-11 on the same rows
            continue

        if current_mineral is None:
            continue

        # Read mining side (col 1 = country, cols 2-5 = years)
        if label_col1 and str(label_col1).strip() not in ("", "Notes"):
            country = str(label_col1).strip()
            if country == "Total":
                vals = {}
                for yr, col in SUPPLY_MINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["mining"]["total"] = vals
            elif country == "Top 3 share":
                vals = {}
                for yr, col in SUPPLY_MINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["mining"]["top3Share"] = vals
            elif country != "Rest of world":
                vals = {}
                for yr, col in SUPPLY_MINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["mining"]["countries"][country] = vals
            else:
                vals = {}
                for yr, col in SUPPLY_MINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["mining"]["countries"]["Rest of world"] = vals

        # Read refining side (col 7 = country, cols 8-11 = years)
        if label_col7 and str(label_col7).strip() not in ("", "Notes"):
            country = str(label_col7).strip()
            if country == "Total" or country == "Total clean technologies":
                vals = {}
                for yr, col in SUPPLY_REFINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["refining"]["total"] = vals
            elif country == "Top 3 share":
                vals = {}
                for yr, col in SUPPLY_REFINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["refining"]["top3Share"] = vals
            elif country != "Rest of world":
                vals = {}
                for yr, col in SUPPLY_REFINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["refining"]["countries"][country] = vals
            else:
                vals = {}
                for yr, col in SUPPLY_REFINING_COLS.items():
                    vals[str(yr)] = safe_num(ws.cell(r, col).value)
                result[current_mineral]["refining"]["countries"]["Rest of world"] = vals

    return result


def parse_sheet32_cleantech_by_mineral(wb):
    """Parse Sheet 3.2: Cleantech demand by mineral (37 minerals)."""
    ws = wb["3.2 Cleantech demand by mineral"]
    result = {}

    skip = {"Mineral demand for clean energy technologies - by mineral (kt)", "Total"}
    for r in range(7, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() in skip:
            continue
        label = str(label).strip()
        if label.startswith("Note"):
            continue
        name = clean_name(label)
        result[name] = read_scenario_row(ws, r)

    return result


def parse_sheet31_cleantech_by_tech(wb):
    """Parse Sheet 3.1: Cleantech demand by tech and mineral."""
    ws = wb["3.1 Cleantech demand by tech"]
    result = {}

    current_mineral = None
    for r in range(7, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()
        if not label:
            continue

        # Check if it's a mineral header (no data in col 2)
        val_b = ws.cell(r, 2).value
        if val_b is None and not label.startswith("Total") and not label.startswith("Note"):
            current_mineral = clean_name(label)
            if current_mineral not in result:
                result[current_mineral] = {"sectors": {}, "total": None}
            continue

        if current_mineral is None:
            continue

        if label.startswith("Total"):
            result[current_mineral]["total"] = read_scenario_row(ws, r)
        elif not label.startswith("Note"):
            result[current_mineral]["sectors"][label] = read_scenario_row(ws, r)

    return result


def parse_tech_sheets(wb):
    """Parse Sheets 4.1-4.6: Per-technology mineral breakdowns."""
    result = {}

    for sheet_name, tech_name in TECH_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        minerals = {}

        # Find "Base case" row, then read minerals below
        in_base_case = False
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label is None:
                continue
            label = str(label).strip()

            if label == "Base case":
                in_base_case = True
                continue

            if not in_base_case:
                continue

            # Stop at next section or blank
            if label.startswith("Wider") or label.startswith("Note") or label.startswith("Innovation"):
                break

            val_b = ws.cell(r, 2).value
            if val_b is None:
                # Could be a sub-header, skip
                continue

            name = clean_name(label)
            minerals[name] = read_scenario_row(ws, r)

        result[tech_name] = {"minerals": minerals}

    return result


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    print(f"Reading {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("  Parsing Sheet 1: Total demand...")
    total_demand = parse_sheet1_total_demand(wb)

    print("  Parsing Sheet 2: Total supply...")
    supply = parse_sheet2_supply(wb)

    print("  Parsing Sheet 3.1: Cleantech by tech...")
    cleantech_by_tech = parse_sheet31_cleantech_by_tech(wb)

    print("  Parsing Sheet 3.2: Cleantech by mineral...")
    cleantech_by_mineral = parse_sheet32_cleantech_by_mineral(wb)

    print("  Parsing Sheets 4.1-4.6: Technology breakdowns...")
    by_technology = parse_tech_sheets(wb)

    output = {
        "source": "IEA Critical Minerals Data Explorer (2024), CC BY 4.0",
        "totalDemand": total_demand,
        "supply": supply,
        "cleantechByTech": cleantech_by_tech,
        "cleantechByMineral": cleantech_by_mineral,
        "byTechnology": by_technology,
    }

    OUT_DIR.mkdir(exist_ok=True)
    out_path = OUT_DIR / "iea.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    size_kb = out_path.stat().st_size / 1024
    print(f"\nWrote {out_path} ({size_kb:.1f} KB)")

    # Summary
    print(f"  totalDemand: {len(total_demand)} minerals")
    print(f"  supply: {len(supply)} minerals")
    print(f"  cleantechByTech: {len(cleantech_by_tech)} minerals")
    print(f"  cleantechByMineral: {len(cleantech_by_mineral)} minerals")
    print(f"  byTechnology: {len(by_technology)} technologies")


if __name__ == "__main__":
    main()
